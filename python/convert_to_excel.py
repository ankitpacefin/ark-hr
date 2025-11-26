
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def flatten_json(y):
    """
    Flattens a nested JSON object.
    """
    out = {}

    def flatten(x, name=''):
        if type(x) is dict:
            for a in x:
                flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out

def process_application_data(data):
    """
    Processes the application data to be suitable for Excel.
    This involves flattening the nested 'ai_data' and formatting lists.
    """
    processed_data = []
    for item in data:
        # Handle cases where ai_data might be missing or there's an error
        if 'ai_data' not in item or 'analysis_error' in item:
            flat_item = item.copy()
            if 'ai_data' in flat_item:
                del flat_item['ai_data'] # remove to avoid partial processing
            processed_data.append(flat_item)
            continue

        ai_data = item.pop('ai_data')

        # Flatten the main application details
        flat_item = item.copy()

        # Process and flatten ai_data separately
        if ai_data:
            flat_item['ai_ats_score'] = ai_data.get('ats_score')
            
            social_links = ai_data.get('social_links', [])
            flat_item['ai_social_links'] = "\\n".join([link for link in social_links if link]) if social_links else ""
            
            flat_item['ai_current_job_title'] = ai_data.get('current_job_title')
            flat_item['ai_gender'] = ai_data.get('gender')
            flat_item['ai_total_experience_years'] = ai_data.get('total_experience_years')
            flat_item['ai_highest_qualification'] = ai_data.get('highest_qualification')

            skills = ai_data.get('skills', [])
            flat_item['ai_skills'] = ", ".join([skill for skill in skills if skill]) if skills else ""

            domains_worked = ai_data.get('domains_worked', [])
            flat_item['ai_domains_worked'] = ", ".join([domain for domain in domains_worked if domain]) if domains_worked else ""

            flat_item['ai_notable_achievement'] = ai_data.get('notable_achievement')
            
            # Handle both 'prevous_companies_names' and 'previous_companies_names'
            prev_companies = ai_data.get('prevous_companies_names', ai_data.get('previous_companies_names', []))
            flat_item['ai_previous_companies'] = ", ".join([company for company in prev_companies if company]) if prev_companies else ""


            projects = ai_data.get('projects', [])
            if projects:
                project_texts = []
                for p in projects:
                    name = p.get('name', 'N/A')
                    desc = p.get('description', 'N/A')
                    url = p.get('url', 'N/A')
                    project_texts.append(f"Name: {name}\\nDescription: {desc}\\nURL: {url}")
                flat_item['ai_projects'] = "\\n---\\n".join(project_texts)
            else:
                flat_item['ai_projects'] = ""

        processed_data.append(flat_item)

    return processed_data


def convert_json_to_excel(json_file_path, excel_file_path):
    """
    Reads a JSON file, processes it, and saves it as a formatted Excel file.
    """
    # Step 1: Install pandas if you don't have it
    # pip install pandas openpyxl

    # Step 2: Read the JSON file
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Step 3: Process the data
    processed_data = process_application_data(data)

    # Step 4: Create a pandas DataFrame
    df = pd.DataFrame(processed_data)

    # Step 5: Save to Excel with formatting using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply formatting to header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Adjust column widths and apply text wrapping and borders
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            cell.border = thin_border
            # Apply text wrapping for multi-line content
            if "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Adjust width, with a cap for very long text
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

    # Freeze the top row
    ws.freeze_panes = 'A2'

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook
    wb.save(excel_file_path)
    print(f"Successfully converted {json_file_path} to {excel_file_path}")


if __name__ == "__main__":
    # Before running, make sure you have pandas and openpyxl installed:
    # pip install pandas openpyxl
    
    JSON_FILE = 'processed_applications.json'
    EXCEL_FILE = 'processed_applications.xlsx'
    convert_json_to_excel(JSON_FILE, EXCEL_FILE)
